# import tkinter as tk
# from tkinter import messagebox
# from blockchain import Blockchain
# from ledger import Ledger
# from block_visualization import visualize_blocks
# from validator import validate_chain 
# import datetime

# class BlockchainApp:
#     def __init__(self, root):
#         self.root = root
#         self.root.title("Blockchain Application")
        
#         self.blockchain = Blockchain()
#         self.ledger = Ledger()

#         self.create_widgets()

#     def create_widgets(self):
#         # Buttons
#         self.add_block_button = tk.Button(self.root, text="Add Block", command=self.add_block)
#         self.add_block_button.pack(pady=10)

#         self.validate_button = tk.Button(self.root, text="Validate Blockchain", command=self.validate_blockchain)
#         self.validate_button.pack(pady=10)
        
#         self.display_button = tk.Button(self.root, text="Display Blockchain", command=self.display_blockchain)
#         self.display_button.pack(pady=10)

#         self.show_ledger_button = tk.Button(self.root, text="Show Ledger", command=self.show_ledger)
#         self.show_ledger_button.pack(pady=10)

#         self.close_button = tk.Button(self.root, text="Close Program", command=self.close_program)
#         self.close_button.pack(pady=10)

#     def add_block(self):
#         data = "Some Data"  # Get data from user or use some predefined data
#         data_type = type(data).__name__

#         self.blockchain.add_block(data)
#         self.ledger.add_record(datetime.datetime.now(), data_type)

#         messagebox.showinfo("Block Added", "A new block has been added!")

#     def validate_blockchain(self):
#         is_valid = self.blockchain.validate_chain()
#         if is_valid:
#             messagebox.showinfo("Blockchain Validation", "Blockchain is valid!")
#         else:
#             messagebox.showerror("Blockchain Validation", "Blockchain is invalid!")
    
#     def display_blockchain(self):
#         return 

#     def show_ledger(self):
#         ledger_data = self.ledger.get_ledger()
#         messagebox.showinfo("Ledger", str(ledger_data))

#     def close_program(self):
#         self.ledger.save_to_excel()
#         self.root.quit()


# import tkinter as tk
# from tkinter import messagebox, simpledialog
# import datetime
# import hashlib
# from blockchain import Blockchain
# from ledger import Ledger
# from validator import validate_chain

# class Block:
#     def __init__(self, data, previous_hash):
#         self.timestamp = datetime.datetime.now()
#         self.data = data
#         self.previous_hash = previous_hash
#         self.hash = self.calculate_hash()

#     def calculate_hash(self):
#         block_string = f"{self.timestamp}{self.data}{self.previous_hash}"
#         return hashlib.sha256(block_string.encode('utf-8')).hexdigest()


# class Blockchain:
#     def __init__(self):
#         self.chain = []  # List to store blocks
#         self.create_genesis_block()

#     def create_genesis_block(self):
#         # Create the first block with arbitrary data and no previous hash
#         genesis_block = Block("Genesis Block", "0")
#         self.chain.append(genesis_block)

#     def add_block(self, data):
#         # Get the last block's hash to link this new block
#         last_block = self.get_last_block()
#         previous_hash = last_block.hash

#         # Add a new block to the blockchain with previous hash
#         new_block = Block(data, previous_hash)
#         self.chain.append(new_block)

#     def get_last_block(self):
#         return self.chain[-1]

#     def validate_chain(self):
#         # Blockchain validation logic (simple check)
#         for i in range(1, len(self.chain)):
#             current_block = self.chain[i]
#             previous_block = self.chain[i - 1]

#             # Check if the hash of the current block matches the stored hash
#             if current_block.previous_hash != previous_block.hash:
#                 return False
#             if current_block.hash != current_block.calculate_hash():
#                 return False
#         return True


# class Ledger:
#     def __init__(self):
#         self.records = []

#     def add_record(self, timestamp, data_type):
#         self.records.append((timestamp, data_type))

#     def get_ledger(self):
#         return self.records

#     def save_to_excel(self):
#         # Placeholder for saving to an Excel file
#         print("Saving ledger to Excel... (Not implemented)")


# class BlockchainApp:
#     def __init__(self, root):
#         self.root = root
#         self.root.title("Blockchain Application")
        
#         self.blockchain = Blockchain()
#         self.ledger = Ledger()

#         self.create_widgets()

#     def create_widgets(self):
#         # Buttons
#         self.add_block_button = tk.Button(self.root, text="Add Block", command=self.add_block)
#         self.add_block_button.pack(pady=10)

#         self.validate_button = tk.Button(self.root, text="Validate Blockchain", command=self.validate_blockchain)
#         self.validate_button.pack(pady=10)
        
#         self.display_button = tk.Button(self.root, text="Display Blockchain", command=self.display_blockchain)
#         self.display_button.pack(pady=10)

#         self.show_ledger_button = tk.Button(self.root, text="Show Ledger", command=self.show_ledger)
#         self.show_ledger_button.pack(pady=10)

#         self.close_button = tk.Button(self.root, text="Close Program", command=self.close_program)
#         self.close_button.pack(pady=10)

#     def add_block(self):
#         # Create a simple Tkinter window to get data from user
#         data = simpledialog.askstring("Input", "Enter data for the new block:", parent=self.root)

#         if data:
#             # Get the last block's hash to link this new block
#             last_block = self.blockchain.get_last_block()
#             previous_hash = last_block.hash

#             # Determine the type of data for the ledger
#             data_type = type(data).__name__

#             # Add the new block to the blockchain
#             self.blockchain.add_block(data)

#             # Add the record to the ledger
#             self.ledger.add_record(datetime.datetime.now(), data_type)

#             messagebox.showinfo("Block Added", "A new block has been added!")
#         else:
#             messagebox.showwarning("Input Error", "No data entered. Block not added.")

#     def validate_blockchain(self):
#         is_valid = self.blockchain.validate_chain()
#         if is_valid:
#             messagebox.showinfo("Blockchain Validation", "Blockchain is valid!")
#         else:
#             messagebox.showerror("Blockchain Validation", "Blockchain is invalid!")

#     def display_blockchain(self):
#         # Display the blockchain (here we just print the hash values)
#         blockchain_data = ""
#         for block in self.blockchain.chain:
#             blockchain_data += f"Timestamp: {block.timestamp}, Data: {block.data}, Hash: {block.hash}\n"
        
#         messagebox.showinfo("Blockchain", blockchain_data)

#     def show_ledger(self):
#         # Show the ledger data
#         ledger_data = self.ledger.get_ledger()
#         ledger_str = "\n".join([f"{record[0]} - {record[1]}" for record in ledger_data])
#         messagebox.showinfo("Ledger", ledger_str)

#     def close_program(self):
#         # Save to Excel before closing
#         self.ledger.save_to_excel()
#         self.root.quit()


# # Main execution
# if __name__ == "__main__":
#     root = tk.Tk()
#     app = BlockchainApp(root)
#     root.mainloop()

import tkinter as tk
from tkinter import messagebox, simpledialog
import datetime
import hashlib
import random
import networkx as nx
import matplotlib.pyplot as plt
from blockchain import Blockchain
from ledger import Ledger
from validator import validate_chain
import tkinter as tk
from tkinter import messagebox, simpledialog
import datetime
from tkcalendar import Calendar
import os
import pandas as pd
from openpyxl import load_workbook


class Block:
    def __init__(self, data, previous_hash):
        self.timestamp = datetime.datetime.now()
        self.data = data
        self.previous_hash = previous_hash
        self.nonce = random.randint(1000, 9999)  # Random nonce generation
        self.hash = self.calculate_hash()

    def calculate_hash(self):
        block_string = f"{self.timestamp}{self.data}{self.previous_hash}{self.nonce}"
        return hashlib.sha256(block_string.encode('utf-8')).hexdigest()


class Blockchain:
    def __init__(self):
        self.chain = []  # List to store blocks
        self.create_genesis_block()

    def create_genesis_block(self):
        # Create the first block with arbitrary data and no previous hash
        genesis_block = Block("Genesis Block", "0")
        self.chain.append(genesis_block)

    def add_block(self, data):
        # Get the last block's hash to link this new block
        last_block = self.get_last_block()
        previous_hash = last_block.hash

        # Add a new block to the blockchain with previous hash
        new_block = Block(data, previous_hash)
        self.chain.append(new_block)

    def get_last_block(self):
        return self.chain[-1]

    def validate_chain(self):
        # Blockchain validation logic (simple check)
        for i in range(1, len(self.chain)):
            current_block = self.chain[i]
            previous_block = self.chain[i - 1]

            # Check if the hash of the current block matches the stored hash
            if current_block.previous_hash != previous_block.hash:
                return False
            if current_block.hash != current_block.calculate_hash():
                return False
        return True


class Ledger:
    def __init__(self, excel_file="ledger.xlsx"):
        self.records = []
        self.excel_file = excel_file

    def add_record(self, timestamp, data_type, previous_hash, data, current_hash, nonce):
        self.records.append((timestamp, data_type, previous_hash, data, current_hash, nonce))

    def get_ledger(self):
        return self.records

    def get_ledger_by_time(self, start_time, end_time):
        filtered_records = [
            record for record in self.records if start_time <= record[0] <= end_time
        ]
        return filtered_records

    def save_to_excel(self):
        if not os.path.exists(self.excel_file):
            # If the Excel file does not exist, create it with a sheet for each day
            with pd.ExcelWriter(self.excel_file, engine='openpyxl') as writer:
                for record in self.records:
                    date = record[0].date()  # Extract the date part of the timestamp
                    sheet_name = str(date)
                    df = pd.DataFrame([record], columns=["Timestamp", "Data Type", "Previous Hash", "Data", "Current Hash", "Nonce"])
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            # If the file exists, update the relevant sheet
            workbook = load_workbook(self.excel_file)
            for record in self.records:
                date = record[0].date()
                sheet_name = str(date)
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                else:
                    sheet = workbook.create_sheet(sheet_name)
                
                df = pd.DataFrame([record], columns=["Timestamp", "Data Type", "Previous Hash", "Data", "Current Hash", "Nonce"])
                for row in dataframe_to_rows(df, index=False, header=False):
                    sheet.append(row)
            workbook.save(self.excel_file)


class BlockchainApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Blockchain Application")
        
        self.blockchain = Blockchain()
        self.ledger = Ledger()

        self.create_widgets()

    def create_widgets(self):
        # Buttons
        self.add_block_button = tk.Button(self.root, text="Add Block", command=self.add_block)
        self.add_block_button.pack(pady=10)

        self.validate_button = tk.Button(self.root, text="Validate Blockchain", command=self.validate_blockchain)
        self.validate_button.pack(pady=10)
        
        self.display_button = tk.Button(self.root, text="Display Blockchain", command=self.display_blockchain)
        self.display_button.pack(pady=10)

        self.show_ledger_button = tk.Button(self.root, text="Show Ledger", command=self.show_ledger)
        self.show_ledger_button.pack(pady=10)

        self.close_button = tk.Button(self.root, text="Close Program", command=self.close_program)
        self.close_button.pack(pady=10)

    def add_block(self):
        # Create a simple Tkinter window to get data from user
        data = simpledialog.askstring("Input", "Enter data for the new block:", parent=self.root)

        if data:
            # Get the last block's hash to link this new block
            last_block = self.blockchain.get_last_block()
            previous_hash = last_block.hash

            # Generate nonce (random number) and current hash
            nonce = random.randint(1000, 9999)

            # Add the new block to the blockchain
            self.blockchain.add_block(data)

            # Add the record to the ledger
            current_block = self.blockchain.get_last_block()
            self.ledger.add_record(
                current_block.timestamp, 
                type(data).__name__, 
                previous_hash, 
                data, 
                current_block.hash, 
                nonce
            )

            messagebox.showinfo("Block Added", "A new block has been added!")
        else:
            messagebox.showwarning("Input Error", "No data entered. Block not added.")

    def validate_blockchain(self):
        is_valid = self.blockchain.validate_chain()
        if is_valid:
            messagebox.showinfo("Blockchain Validation", "Blockchain is valid!")
        else:
            messagebox.showerror("Blockchain Validation", "Blockchain is invalid!")

    def display_blockchain(self):
        # Display the blockchain as a graph
        G = nx.DiGraph()  # Directed graph for visualizing the chain
        for i, block in enumerate(self.blockchain.chain):
            G.add_node(
                i,
                label=f"Block {i+1}\nData: {block.data}\nTimestamp: {block.timestamp}\nHash: {block.hash[:6]}..."
            )
            if i > 0:
                G.add_edge(i-1, i)

        pos = nx.spring_layout(G)
        labels = nx.get_node_attributes(G, "label")
        nx.draw(G, pos, with_labels=True, node_size=3000, node_color="skyblue", font_size=10)
        nx.draw_networkx_labels(G, pos, labels, font_size=8)
        plt.show()

    def show_ledger(self):
        # Ask the user for the date to filter the ledger
        def on_date_select():
            selected_date = calendar.get_date()
            # Convert the selected date to a datetime object
            selected_date = datetime.datetime.strptime(selected_date, "%m/%d/%Y").date()

            # Filter ledger for the selected date
            filtered_ledger = [record for record in self.ledger.get_ledger() if record[0].date() == selected_date]

            if filtered_ledger:
                ledger_str = "\n".join([f"{record[0]} - {record[1]} - {record[2]} - {record[3]} - {record[4]} - {record[5]}" for record in filtered_ledger])
                messagebox.showinfo("Ledger", ledger_str)
            else:
                messagebox.showinfo("No Records", "No records found for the selected date.")

            # Destroy the calendar popup
            calendar_popup.destroy()

        # Create a pop-up window for date selection
        calendar_popup = tk.Toplevel(self.root)
        calendar_popup.title("Select Date")

        # Create a Calendar widget for date selection
        calendar = Calendar(calendar_popup, selectmode="day", date_pattern="mm/dd/yyyy")
        calendar.pack(pady=20)

        # Add a button to confirm the date selection
        select_button = tk.Button(calendar_popup, text="Select Date", command=on_date_select)
        select_button.pack(pady=10)

        # Wait until the user selects a date and closes the popup
        calendar_popup.mainloop()

    def close_program(self):
        # Save to Excel before closing
        self.ledger.save_to_excel()
        self.root.quit()


# Main execution
if __name__ == "__main__":
    root = tk.Tk()
    app = BlockchainApp(root)
    root.mainloop()

